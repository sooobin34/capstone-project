import os
import urllib.request
from collections import Counter, defaultdict
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from math import ceil

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font as XLFont, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from app.api.deps import get_db
from app.models.awd_daily_summary import AwdDailySummary
from app.models.field import Field
from app.models.iot_node import IotNode
from app.models.mrv_report import MrvReport
from app.models.validation_record import ValidationRecord
from app.schemas.mrv_report import MrvReportCreate, MrvReportStatusUpdate
from app.utils.response import success_response
from app.models.sensor_log import SensorLog

router = APIRouter(prefix="/mrv-reports", tags=["MRV Reports"])

ALLOWED_MRV_STATUSES = {"IN_PROGRESS", "COMPLETED"}

COVER_TITLE_SIZE = 20
COVER_SUBTITLE_SIZE = 15
COVER_INFO_SIZE = 10
TOC_TITLE_SIZE = 15
TOC_BODY_SIZE = 11
SECTION_TITLE_SIZE = 12
SUB_TITLE_SIZE = 11
BODY_SIZE = 10
BODY_LINE_HEIGHT = 15
PARAGRAPH_GAP = 9
SECTION_GAP = 18
LEFT_X = 45
RIGHT_MARGIN = 45
TOP_Y = 756
BOTTOM_Y = 55


def _font_path(filename: str) -> str:
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "fonts", filename))


def register_korean_fonts() -> tuple[str, str]:
    regular_candidates = [
        _font_path("NanumMyeongjo.ttf"),
        _font_path("NanumGothic.ttf"),
        r"C:\Windows\Fonts\malgun.ttf",
    ]
    bold_candidates = [
        _font_path("NanumMyeongjoBold.ttf"),
        _font_path("NanumGothicBold.ttf"),
        r"C:\Windows\Fonts\malgunbd.ttf",
    ]

    regular_font = None
    bold_font = None

    for path in regular_candidates:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont("KoreanRegular", path))
            regular_font = "KoreanRegular"
            break

    for path in bold_candidates:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont("KoreanBold", path))
            bold_font = "KoreanBold"
            break

    if regular_font and not bold_font:
        bold_font = regular_font

    if regular_font and bold_font:
        return regular_font, bold_font

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
        return "HYSMyeongJo-Medium", "HYSMyeongJo-Medium"
    except Exception:
        raise RuntimeError(
            "사용 가능한 한글 폰트를 찾지 못했습니다. "
            "backend/app/fonts 폴더에 NanumGothic.ttf, NanumMyeongjo.ttf, NanumMyeongjoBold.ttf를 추가하세요."
        )


def get_month_range(report_month: str) -> tuple[date, date]:
    year, month = map(int, report_month.split("-"))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    return start_date, end_date


def format_report_month(report_month: str) -> str:
    year, month = map(int, report_month.split("-"))
    return f"{year}년 {month}월"


def select_representative_validation_rows(rows: list[ValidationRecord]) -> list[tuple[str, ValidationRecord]]:
    valid_rows = [row for row in rows if row.image_url]
    if not valid_rows:
        return []

    valid_rows = sorted(valid_rows, key=lambda x: (x.record_date, x.id))
    n = len(valid_rows)
    labels = ["월 초", "월 중", "월 말"]
    raw_indices = [0, n // 2, n - 1]

    selected = []
    seen_urls = set()
    for label, idx in zip(labels, raw_indices):
        row = valid_rows[idx]
        if row.image_url not in seen_urls:
            selected.append((label, row))
            seen_urls.add(row.image_url)

    return selected


def get_validation_rows(field_id: int, start_date: date, end_date: date, db: Session) -> list[ValidationRecord]:
    return (
        db.query(ValidationRecord)
        .filter(
            ValidationRecord.field_id == field_id,
            ValidationRecord.record_date >= start_date,
            ValidationRecord.record_date < end_date,
        )
        .order_by(ValidationRecord.record_date.asc(), ValidationRecord.id.asc())
        .all()
    )


def get_validation_summary(field_id: int, start_date: date, end_date: date, db: Session) -> dict:
    rows = get_validation_rows(field_id, start_date, end_date, db)

    sample_count = len(rows)

    sensor_valid_rows = [row for row in rows if row.is_match is not None]
    match_count = sum(1 for row in sensor_valid_rows if row.is_match is True)
    mismatch_count = sum(1 for row in sensor_valid_rows if row.is_match is False)
    unknown_count = sum(1 for row in rows if row.is_match is None)
    accuracy = round((match_count / len(sensor_valid_rows)) * 100, 2) if sensor_valid_rows else 0

    ai_sensor_rows = [row for row in rows if row.ai_sensor_match is not None]
    ai_sensor_match_count = sum(1 for row in ai_sensor_rows if row.ai_sensor_match is True)
    ai_sensor_mismatch_count = sum(1 for row in ai_sensor_rows if row.ai_sensor_match is False)
    ai_sensor_unknown_count = sample_count - len(ai_sensor_rows)
    ai_sensor_accuracy = round((ai_sensor_match_count / len(ai_sensor_rows)) * 100, 2) if ai_sensor_rows else 0

    notes = [row.note for row in rows if row.note]
    note = notes[-1] if notes else "별도 비고 없음"

    representative_rows = select_representative_validation_rows(rows)

    return {
        "validation_method": "현장 사진 비교" if sample_count > 0 else "검증 데이터 없음",
        "validation_sample_count": sample_count,
        "validation_match_count": match_count,
        "validation_mismatch_count": mismatch_count,
        "validation_unknown_count": unknown_count,
        "validation_accuracy": accuracy,
        "ai_sensor_match_count": ai_sensor_match_count,
        "ai_sensor_mismatch_count": ai_sensor_mismatch_count,
        "ai_sensor_unknown_count": ai_sensor_unknown_count,
        "ai_sensor_accuracy": ai_sensor_accuracy,
        "validation_note": note,
        "representative_rows": representative_rows,
        "representative_images": [row.image_url for _, row in representative_rows],
    }


def summarize_status_counts(summaries: list[AwdDailySummary]) -> dict:
    counter = Counter(summary.daily_status for summary in summaries)
    return {
        "OVERFLOODED": counter.get("OVERFLOODED", 0),
        "FLOODED": counter.get("FLOODED", 0),
        "DRYING": counter.get("DRYING", 0),
        "DRY": counter.get("DRY", 0),
    }


def aggregate_daily_summaries(summaries: list[AwdDailySummary]) -> list[AwdDailySummary]:
    grouped = defaultdict(list)
    for s in summaries:
        grouped[s.record_date].append(s)

    aggregated = []
    for record_date, items in grouped.items():
        avg_values = [float(i.avg_inner_level) for i in items if i.avg_inner_level is not None]
        if not avg_values:
            continue

        avg_level = sum(avg_values) / len(avg_values)
        if avg_level >= 5:
            status = "OVERFLOODED"
        elif avg_level >= 0:
            status = "FLOODED"
        elif avg_level > -15:
            status = "DRYING"
        else:
            status = "DRY"

        aggregated.append(
            AwdDailySummary(
                record_date=record_date,
                avg_inner_level=avg_level,
                daily_status=status,
            )
        )

    return sorted(aggregated, key=lambda x: x.record_date)


def group_summaries_by_week(summaries: list[AwdDailySummary]) -> list[tuple[int, list[AwdDailySummary]]]:
    summaries = sorted(summaries, key=lambda x: x.record_date)
    grouped: dict[int, list[AwdDailySummary]] = defaultdict(list)
    for summary in summaries:
        week_no = ceil(summary.record_date.day / 7)
        grouped[week_no].append(summary)
    return sorted(grouped.items(), key=lambda x: x[0])


def get_status_flow(summaries: list[AwdDailySummary]) -> str:
    statuses = [s.daily_status for s in summaries if s.daily_status]
    return " → ".join(statuses) if statuses else "데이터 없음"


def count_awd_cycles(summaries: list[AwdDailySummary]) -> int:
    ordered = sorted(summaries, key=lambda x: x.record_date)
    statuses = [s.daily_status for s in ordered if s.daily_status]
    count = 0
    for prev_status, current_status in zip(statuses, statuses[1:]):
        if prev_status == "DRY" and current_status in ("DRYING", "FLOODED", "OVERFLOODED"):
            count += 1
    return count


def get_month_avg_inner_level(summaries: list[AwdDailySummary]) -> float | None:
    values = [float(summary.avg_inner_level) for summary in summaries if summary.avg_inner_level is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def get_dominant_status(status_counts: dict) -> str:
    if not status_counts:
        return "데이터 없음"
    return max(status_counts.items(), key=lambda x: x[1])[0]


def make_weekly_summary_text(week_no: int, summaries: list[AwdDailySummary]) -> list[str]:
    if not summaries:
        return [
            f"{week_no}주차",
            "해당 주차에는 수집된 데이터가 없어 수위 변화 분석이 제한됩니다.",
        ]

    ordered = sorted(summaries, key=lambda x: x.record_date)
    statuses = [s.daily_status for s in ordered if s.daily_status]
    avg_values = [float(s.avg_inner_level) for s in ordered if s.avg_inner_level is not None]

    if not statuses or not avg_values:
        return [
            f"{week_no}주차",
            "해당 주차에는 상태 또는 평균 수위 데이터가 부족하여 수위 변화 분석이 제한됩니다.",
        ]

    start_avg = avg_values[0]
    end_avg = avg_values[-1]
    min_avg = min(avg_values)
    max_avg = max(avg_values)
    start_status = statuses[0]
    end_status = statuses[-1]

    lines = [
        f"{week_no}주차",
        f"시작 평균 수위: {start_avg:.2f}cm / 최저: {min_avg:.2f}cm / 최고: {max_avg:.2f}cm / 마지막: {end_avg:.2f}cm",
        f"상태 변화: {get_status_flow(ordered)}",
    ]

    if end_avg > start_avg:
        lines.append(f"{week_no}주차에는 평균 내부 수위가 {start_avg:.2f}cm에서 {end_avg:.2f}cm로 상승하였다.")
    elif end_avg < start_avg:
        lines.append(f"{week_no}주차에는 평균 내부 수위가 {start_avg:.2f}cm에서 {end_avg:.2f}cm로 감소하였다.")
    else:
        lines.append(f"{week_no}주차에는 평균 내부 수위가 {start_avg:.2f}cm 수준에서 큰 변화 없이 유지되었다.")

    if "OVERFLOODED" in statuses:
        lines.append("일부 구간에서 OVERFLOODED 상태가 확인되었다.")
    elif "DRY" in statuses:
        lines.append("DRY 상태가 관측된 구간이 확인되었다.")
    elif "DRYING" in statuses:
        lines.append("DRYING 상태가 관측되었다.")

    lines.append(f"해당 주차는 {start_status} 상태에서 시작하여 {end_status} 상태로 마무리되었다.")
    return lines



def draw_page_frame(pdf, width, height):
    pdf.setStrokeColor(colors.lightgrey)
    pdf.setLineWidth(0.3)
    pdf.line(35, height - 35, 55, height - 35)
    pdf.line(35, height - 35, 35, height - 55)
    pdf.line(width - 35, height - 35, width - 55, height - 35)
    pdf.line(width - 35, height - 35, width - 35, height - 55)



def draw_page_number(pdf, width, page_no: int, regular_font: str):
    pdf.setFont(regular_font, 9)
    pdf.setFillColor(colors.black)
    pdf.drawCentredString(width / 2, 28, str(page_no))

def wrap_text(pdf, text: str, max_width: int, font_name: str, font_size: int) -> list[str]:
    if text is None:
        return []
    paragraphs = str(text).split("\n")
    wrapped = []
    for paragraph in paragraphs:
        if paragraph == "":
            wrapped.append("")
            continue
        words = paragraph.split()
        line = ""
        for word in words:
            test_line = f"{line} {word}".strip()
            if pdf.stringWidth(test_line, font_name, font_size) <= max_width:
                line = test_line
            else:
                if line:
                    wrapped.append(line)
                line = word
        if line:
            wrapped.append(line)
    return wrapped


def draw_text(pdf, text: str, x: int, y: int, max_width: int, font_name: str, font_size: int = BODY_SIZE,
              line_height: int = BODY_LINE_HEIGHT, paragraph_gap: int = 0) -> int:
    pdf.setFont(font_name, font_size)
    lines = wrap_text(pdf, text, max_width, font_name, font_size)
    for line in lines:
        if line == "":
            y -= paragraph_gap or line_height
        else:
            pdf.drawString(x, y, line)
            y -= line_height
    return y


def draw_bullets(pdf, lines: list[str], x: int, y: int, max_width: int, font_name: str, font_size: int = BODY_SIZE) -> int:
    pdf.setFont(font_name, font_size)
    for line in lines:
        y = draw_text(pdf, f"- {line}", x, y, max_width, font_name, font_size, BODY_LINE_HEIGHT)
    return y


def draw_section_title(pdf, title: str, y: int, regular_font: str, bold_font: str) -> int:
    pdf.setFont(bold_font, SECTION_TITLE_SIZE)
    pdf.setFillColor(colors.black)
    pdf.drawString(LEFT_X, y, title)
    y -= 24
    return y


def draw_sub_title(pdf, title: str, y: int, bold_font: str) -> int:
    pdf.setFont(bold_font, SUB_TITLE_SIZE)
    pdf.drawString(LEFT_X, y, title)
    y -= 20
    return y


def draw_divider(pdf, y: int, width: int) -> int:
    pdf.setStrokeColor(colors.black)
    pdf.setLineWidth(0.8)
    pdf.line(LEFT_X, y, width - RIGHT_MARGIN, y)
    return y - 26


def draw_simple_table(pdf, x: int, y: int, headers: list[str], rows: list[list[str]], col_widths: list[int],
                      regular_font: str, bold_font: str) -> int:
    row_height = 22
    table_width = sum(col_widths)

    pdf.setFillColor(colors.HexColor("#E9EEF4"))
    pdf.rect(x, y - row_height + 5, table_width, row_height, fill=True, stroke=False)
    pdf.setFillColor(colors.black)

    pdf.setStrokeColor(colors.grey)
    pdf.setLineWidth(0.4)

    current_x = x
    pdf.setFont(bold_font, BODY_SIZE)
    for i, header in enumerate(headers):
        pdf.drawString(current_x + 6, y - 10, header)
        current_x += col_widths[i]

    pdf.rect(x, y - row_height + 5, table_width, row_height, fill=False, stroke=True)
    current_x = x
    for w in col_widths[:-1]:
        current_x += w
        pdf.line(current_x, y + 5, current_x, y - row_height + 5)

    y -= row_height

    pdf.setFont(regular_font, BODY_SIZE)
    for row in rows:
        current_x = x
        for i, value in enumerate(row):
            pdf.drawString(current_x + 6, y - 10, str(value))
            current_x += col_widths[i]
        pdf.rect(x, y - row_height + 5, table_width, row_height, fill=False, stroke=True)
        current_x = x
        for w in col_widths[:-1]:
            current_x += w
            pdf.line(current_x, y + 5, current_x, y - row_height + 5)
        y -= row_height

    return y - 12


def collapse_status_flow(summaries: list) -> str:
    ordered = sorted(summaries, key=lambda s: s.record_date)
    statuses = [s.daily_status for s in ordered if s.daily_status]
    collapsed: list[str] = []
    for s in statuses:
        if not collapsed or collapsed[-1] != s:
            collapsed.append(s)
    return " → ".join(collapsed) if collapsed else "데이터 없음"


def calc_level_change(summaries: list) -> float | None:
    ordered = sorted(summaries, key=lambda s: s.record_date)
    values = [float(s.avg_inner_level) for s in ordered if s.avg_inner_level is not None]
    if len(values) < 2:
        return None
    return round(values[-1] - values[0], 2)


def build_weekly_table_rows(weekly_dict: dict) -> list[list[str]]:
    rows: list[list[str]] = []
    for week_no in sorted(weekly_dict.keys()):
        items = weekly_dict.get(week_no, [])
        if not items:
            continue
        ordered = sorted(items, key=lambda s: s.record_date)
        values = [float(s.avg_inner_level) for s in ordered if s.avg_inner_level is not None]
        if not values:
            continue
        change = calc_level_change(ordered)
        rows.append([
            f"{week_no}주",
            f"{sum(values) / len(values):.2f}",
            f"{change:+.2f}" if change is not None else "-",
            f"{min(values):.2f}",
            f"{max(values):.2f}",
            _flow_text(collapse_status_flow(ordered)),
        ])
    return rows


def draw_status_bar_chart(pdf, x: int, y: int, status_counts: dict, regular_font: str, bold_font: str) -> int:
    statuses = ["OVERFLOODED", "FLOODED", "DRYING", "DRY"]
    counts = [status_counts.get(s, 0) for s in statuses]
    max_days = max(counts) if counts and max(counts) > 0 else 1
    label_w = 110
    bar_max_w = 280
    row_h = 22
    for status, days in zip(statuses, counts):
        pdf.setFont(regular_font, BODY_SIZE)
        pdf.setFillColor(colors.black)
        pdf.drawString(x, y - 10, status)
        pdf.setFillColor(colors.HexColor("#EAEFF4"))
        pdf.roundRect(x + label_w, y - 14, bar_max_w, 12, 3, fill=True, stroke=False)
        bar_w = bar_max_w * days / max_days
        if bar_w > 0:
            pdf.setFillColor(colors.HexColor("#5B7C99"))
            pdf.roundRect(x + label_w, y - 14, bar_w, 12, 3, fill=True, stroke=False)
        pdf.setFillColor(colors.black)
        pdf.setFont(bold_font, BODY_SIZE)
        pdf.drawString(x + label_w + bar_max_w + 10, y - 11, f"{days}일")
        y -= row_h
    return y - 6


def ensure_space_for_validation(
    pdf,
    y: int,
    needed: int,
    width: int,
    height: int,
    regular_font: str,
    page_no_ref: list[int],
) -> int:
    if y < needed:
        draw_page_number(pdf, width, page_no_ref[0], regular_font)
        pdf.showPage()
        page_no_ref[0] += 1
        pdf.setFont(regular_font, BODY_SIZE)
        return TOP_Y
    return y


def load_image_reader(url_or_path: str) -> ImageReader | None:
    if not url_or_path:
        return None
    try:
        if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
            with urllib.request.urlopen(url_or_path, timeout=8) as response:
                data = response.read()
            return ImageReader(BytesIO(data))
        if os.path.exists(url_or_path):
            return ImageReader(url_or_path)
    except Exception:
        return None
    return None


def validation_result_text(value):
    if value is True:
        return "일치"
    if value is False:
        return "불일치"
    return "판별 불가"


def get_sensor_ai_status_for_validation(row: ValidationRecord, db: Session):
    if row.node_id is None:
        return None

    sensor_level = None

    if row.captured_at is not None:
        target = row.captured_at.replace(tzinfo=None)
        start_at = target - timedelta(hours=3)
        end_at = target + timedelta(hours=3)

        candidates = (
            db.query(SensorLog)
            .filter(
                SensorLog.node_id == row.node_id,
                SensorLog.measured_at >= start_at,
                SensorLog.measured_at <= end_at,
            )
            .all()
        )

        if candidates:
            nearest_log = min(
                candidates,
                key=lambda log: abs((log.measured_at.replace(tzinfo=None) - target).total_seconds()),
            )
            sensor_level = nearest_log.inner_water_level

    if sensor_level is None:
        summary = (
            db.query(AwdDailySummary)
            .filter(
                AwdDailySummary.node_id == row.node_id,
                AwdDailySummary.record_date == row.record_date,
            )
            .order_by(AwdDailySummary.id.desc())
            .first()
        )
        sensor_level = summary.avg_inner_level if summary else None

    if sensor_level is None:
        return None

    value = float(sensor_level)
    if value < 2:
        return "LOW"
    if value < 4:
        return "MID"
    return "HIGH"


def validation_status_text(row: ValidationRecord, db: Session) -> str:
    sensor_observed_result = validation_result_text(row.is_match)
    ai_sensor_result = validation_result_text(row.ai_sensor_match)

    observed = row.observed_surface_status or "관찰값 없음"
    sensor_surface = row.sensor_predicted_status or "센서 표면 판정 없음"
    sensor_ai_status = get_sensor_ai_status_for_validation(row, db) or "센서 수위 구간 없음"
    ai = row.ai_predicted_status or "AI 분석 없음"

    return (
        f"촬영일: {row.record_date} / "
        f"사람 관찰값: {observed} / "
        f"센서 기반 표면 판정: {sensor_surface} / "
        f"센서-관찰 검증: {sensor_observed_result} / "
        f"센서 수위 구간: {sensor_ai_status} / "
        f"AI 예측 구간: {ai} / "
        f"AI-센서 구간 비교: {ai_sensor_result}"
    )



@router.post("")
def create_mrv_report(payload: MrvReportCreate, db: Session = Depends(get_db)):
    field = db.query(Field).filter(Field.id == payload.field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="해당 field_id가 존재하지 않습니다.")

    existing_report = (
        db.query(MrvReport)
        .filter(
            MrvReport.field_id == payload.field_id,
            MrvReport.report_month == payload.report_month,
        )
        .first()
    )
    if existing_report:
        raise HTTPException(status_code=400, detail="해당 월의 MRV 보고서가 이미 존재합니다.")

    start_date, end_date = get_month_range(payload.report_month)
    summaries = (
        db.query(AwdDailySummary)
        .join(IotNode, IotNode.id == AwdDailySummary.node_id)
        .filter(
            IotNode.field_id == payload.field_id,
            AwdDailySummary.record_date >= start_date,
            AwdDailySummary.record_date < end_date,
        )
        .all()
    )

    if not summaries:
        raise HTTPException(status_code=404, detail="해당 월의 일일 요약 데이터가 없습니다.")

    daily_summaries = aggregate_daily_summaries(summaries)
    total_awd_cycles = count_awd_cycles(daily_summaries)
    flood_days = sum(1 for s in daily_summaries if s.daily_status in ("FLOODED", "OVERFLOODED"))
    carbon_reduction = (Decimal(total_awd_cycles) * Decimal("15.25")).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )

    report = MrvReport(
        field_id=payload.field_id,
        report_month=payload.report_month,
        total_awd_cycles=total_awd_cycles,
        flood_days=flood_days,
        status="IN_PROGRESS",
        carbon_reduction=carbon_reduction,
    )

    db.add(report)
    db.commit()
    db.refresh(report)

    return success_response(report, "MRV 보고서 생성 성공")


@router.get("")
def get_mrv_reports(field_id: int | None = Query(default=None), db: Session = Depends(get_db)):
    query = db.query(MrvReport, Field.field_name).join(Field, Field.id == MrvReport.field_id)
    if field_id is not None:
        query = query.filter(MrvReport.field_id == field_id)

    rows = query.order_by(MrvReport.created_at.desc()).all()
    reports = []

    for report, field_name in rows:
        start_date, end_date = get_month_range(report.report_month)
        validation = get_validation_summary(report.field_id, start_date, end_date, db)
        reports.append({
            "id": report.id,
            "field_id": report.field_id,
            "field_name": field_name,
            "report_month": report.report_month,
            "total_awd_cycles": report.total_awd_cycles,
            "flood_days": report.flood_days,
            "status": report.status,
            "carbon_reduction": report.carbon_reduction,
            "validation_method": validation["validation_method"],
            "validation_sample_count": validation["validation_sample_count"],
            "validation_match_count": validation["validation_match_count"],
            "validation_accuracy": validation["validation_accuracy"],
            "ai_sensor_match_count": validation["ai_sensor_match_count"],
            "ai_sensor_mismatch_count": validation["ai_sensor_mismatch_count"],
            "ai_sensor_accuracy": validation["ai_sensor_accuracy"],
            "validation_note": validation["validation_note"],
            "created_at": report.created_at,
        })

    return success_response(reports, "MRV 보고서 조회 성공")


@router.patch("/{report_id}/status")
def update_mrv_report_status(report_id: int, payload: MrvReportStatusUpdate, db: Session = Depends(get_db)):
    if payload.status not in ALLOWED_MRV_STATUSES:
        raise HTTPException(status_code=400, detail="status는 IN_PROGRESS 또는 COMPLETED만 가능합니다.")

    report = db.query(MrvReport).filter(MrvReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="해당 report_id가 존재하지 않습니다.")

    report.status = payload.status
    db.commit()
    db.refresh(report)

    return success_response(report, "MRV 보고서 상태 변경 성공")


@router.get("/{report_id}/view")
def get_mrv_report_view(report_id: int, db: Session = Depends(get_db)):
    report = db.query(MrvReport).filter(MrvReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found.")

    field = db.query(Field).filter(Field.id == report.field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found.")

    start_date, end_date = get_month_range(report.report_month)
    summaries = (
        db.query(AwdDailySummary)
        .join(IotNode, IotNode.id == AwdDailySummary.node_id)
        .filter(
            IotNode.field_id == report.field_id,
            AwdDailySummary.record_date >= start_date,
            AwdDailySummary.record_date < end_date,
        )
        .order_by(AwdDailySummary.record_date.asc(), AwdDailySummary.id.asc())
        .all()
    )
    nodes = db.query(IotNode).filter(IotNode.field_id == report.field_id).order_by(IotNode.id.asc()).all()

    daily_summaries = aggregate_daily_summaries(summaries)
    status_counts = summarize_status_counts(daily_summaries)
    dominant_status = get_dominant_status(status_counts)
    month_avg = get_month_avg_inner_level(daily_summaries)
    weekly_groups = group_summaries_by_week(daily_summaries)

    weekly_analysis = []
    for week_no, week_items in weekly_groups:
        if not week_items:
            continue

        ordered_week_items = sorted(week_items, key=lambda item: item.record_date)
        values = [float(item.avg_inner_level) for item in ordered_week_items if item.avg_inner_level is not None]
        start_level = values[0] if values else None
        end_level = values[-1] if values else None
        weekly_analysis.append(
            {
                "week_no": week_no,
                "start_date": ordered_week_items[0].record_date.isoformat(),
                "end_date": ordered_week_items[-1].record_date.isoformat(),
                "avg_inner_level_cm": round(sum(values) / len(values), 2) if values else None,
                "start_inner_level_cm": round(start_level, 2) if start_level is not None else None,
                "end_inner_level_cm": round(end_level, 2) if end_level is not None else None,
                "change_inner_level_cm": calc_level_change(ordered_week_items),
                "min_inner_level_cm": round(min(values), 2) if values else None,
                "max_inner_level_cm": round(max(values), 2) if values else None,
                "status_flow": collapse_status_flow(ordered_week_items),
            }
        )

    validation = get_validation_summary(report.field_id, start_date, end_date, db)
    validation_rows = get_validation_rows(report.field_id, start_date, end_date, db)
    validation_snapshots = [
        {
            "record_id": row.id,
            "record_date": row.record_date.isoformat(),
            "node_id": row.node_id,
            "sensor_predicted_status": row.sensor_predicted_status,
            "observed_surface_status": row.observed_surface_status,
            "ai_predicted_status": row.ai_predicted_status,
            "ai_confidence": float(row.ai_confidence) if row.ai_confidence is not None else None,
            "sensor_observed_match": row.is_match,
            "ai_sensor_match": row.ai_sensor_match,
            "image_url": row.image_url,
            "note": row.note,
        }
        for row in validation_rows
    ]

    base_conclusion = [
        "본 보고서는 탄소배출권 거래를 최종 목표로 하는 단계적 연구의 초기 설계 단계 결과로, "
        "수위 실측·현장 검증·기록 및 보고 문서화를 통해 MRV 기반 체계를 구축하는 데 중점을 두었다.",
        "핵심 성과는 IoT 수위 실측, 현장 사진 검증, 기록 데이터, 보고 문서화를 하나의 MRV 흐름으로 연결한 점이다.",
    ]
    if report.total_awd_cycles == 0:
        conclusion = base_conclusion + [
            "보고 기간 중 완결된 AWD 사이클은 관측되지 않았으나, 센서 실측과 현장 검증을 통해 "
            "탄소배출권 산정의 근거가 되는 MRV 데이터를 확보하였다.",
            "본 연구는 탄소배출권 플랫폼 기반 구축에 해당하며, 향후 모니터링 기간 확대와 제도 연계를 통해 "
            "배출권 거래 단계까지 확장하고자 한다.",
        ]
    else:
        conclusion = base_conclusion + [
            f"보고 기간 중 {report.total_awd_cycles}회의 AWD 사이클이 관측되었으며, "
            "수위 데이터 수집·현장 검증·기록·보고 문서화 과정을 자동화하였다.",
            "본 보고서는 탄소배출권 거래 또는 공식 탄소감축량 산정을 완료한 결과물이 아니라, "
            "향후 탄소감축량 산정 및 탄소배출권 제도 연계를 위한 MRV 기반 자료로 활용될 수 있다.",
        ]

    payload = {
        "report_id": report.id,
        "overview": {
            "field_id": field.id,
            "field_name": field.field_name,
            "field_location_desc": field.location_desc,
            "report_month": report.report_month,
            "period_start": start_date.isoformat(),
            "period_end_exclusive": end_date.isoformat(),
            "node_count": len(nodes),
            "generated_at": report.created_at.isoformat() if report.created_at else None,
            "status": report.status,
        },
        "summary": {
            "total_awd_cycles": report.total_awd_cycles,
            "flood_days": report.flood_days,
            "carbon_reduction_kgco2eq": float(report.carbon_reduction) if report.carbon_reduction is not None else None,
            "dominant_status": dominant_status,
            "month_avg_inner_level_cm": month_avg,
            "status_counts": status_counts,
        },
        "weekly_analysis": weekly_analysis,
        "validation_results": {
            "validation_method": validation["validation_method"],
            "sample_count": validation["validation_sample_count"],
            "sensor_observed_match_count": validation["validation_match_count"],
            "sensor_observed_mismatch_count": validation["validation_mismatch_count"],
            "sensor_observed_unknown_count": validation["validation_unknown_count"],
            "sensor_observed_accuracy": validation["validation_accuracy"],
            "ai_sensor_match_count": validation["ai_sensor_match_count"],
            "ai_sensor_mismatch_count": validation["ai_sensor_mismatch_count"],
            "ai_sensor_unknown_count": validation["ai_sensor_unknown_count"],
            "ai_sensor_accuracy": validation["ai_sensor_accuracy"],
            "note": validation["validation_note"],
            "rows": validation_snapshots,
        },
        "conclusion": conclusion,
        "download": {
            "pdf_url": f"/mrv-reports/{report.id}/download/pdf",
            "excel_url": f"/mrv-reports/{report.id}/download/excel",
        },
    }

    return success_response(payload, "MRV report view data retrieved successfully.")


def _mrv_view_data(report_id: int, db: Session) -> dict:
    return get_mrv_report_view(report_id, db)["data"]


def _fmt_value(value, suffix: str = "") -> str:
    if value is None:
        return "-"
    if isinstance(value, float):
        return f"{value:.2f}{suffix}"
    return f"{value}{suffix}"


def _status_label(status: str | None) -> str:
    labels = {
        "OVERFLOODED": "OVERFLOODED",
        "FLOODED": "FLOODED",
        "DRYING": "DRYING",
        "DRY": "DRY",
        "NO_DATA": "NO_DATA",
    }
    return labels.get(status or "", status or "-")


def _status_color(status: str | None) -> str:
    return "#1D9E75"


def _match_label(value) -> str:
    if value is True:
        return "일치"
    if value is False:
        return "불일치"
    return "판정 불가"


def _flow_text(value: str | None) -> str:
    if not value:
        return "-"
    normalized = value.replace("->", "→").replace(" ??", "→").replace("??", "→")
    parts = [part.strip() for part in normalized.split("→") if part.strip()]
    if not parts:
        return "-"
    return " → ".join(parts)


def _draw_pdf_header(pdf, width: int, height: int, title: str, regular_font: str, bold_font: str):
    pdf.setFillColor(colors.white)
    pdf.rect(0, height - 58, width, 58, fill=True, stroke=False)
    pdf.setStrokeColor(colors.HexColor("#D9E2E8"))
    pdf.setLineWidth(0.6)
    pdf.line(42, height - 54, width - 42, height - 54)
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 13)
    pdf.drawString(42, height - 34, title)
    pdf.setStrokeColor(colors.HexColor("#0F6B4F"))
    pdf.setLineWidth(1.1)
    pdf.line(42, height - 58, 118, height - 58)
    pdf.setFillColor(colors.HexColor("#667085"))
    pdf.setFont(regular_font, 9)
    pdf.drawRightString(width - 42, height - 34, "AquaPaddy MRV")


def _start_pdf_body_page(pdf, width: int, height: int) -> int:
    draw_page_frame(pdf, width, height)
    return height - 82


def _draw_pdf_footer(pdf, width: int, page_no: int, regular_font: str):
    pdf.setStrokeColor(colors.HexColor("#D9E2E8"))
    pdf.setLineWidth(0.5)
    pdf.line(42, 42, width - 42, 42)
    pdf.setFillColor(colors.HexColor("#667085"))
    pdf.setFont(regular_font, 8)
    pdf.drawString(42, 28, "IoT water-level monitoring based MRV summary")
    pdf.drawRightString(width - 42, 28, f"page {page_no}")


def _draw_pdf_section_label(pdf, x: int, y: int, label: str, bold_font: str) -> int:
    pdf.setFillColor(colors.HexColor("#1F2937"))
    pdf.setFont(bold_font, 12)
    pdf.drawString(x, y - 14, label)
    pdf.setStrokeColor(colors.HexColor("#0F6B4F"))
    pdf.setLineWidth(0.9)
    pdf.line(x, y - 22, x + 120, y - 22)
    return y - 34


def _draw_pdf_card(pdf, x: int, y: int, w: int, h: int, title: str, value: str, regular_font: str, bold_font: str):
    pdf.setFillColor(colors.white)
    pdf.setStrokeColor(colors.HexColor("#D9E2E8"))
    pdf.rect(x, y - h, w, h, fill=True, stroke=True)
    pdf.setFillColor(colors.HexColor("#0F6B4F"))
    pdf.rect(x, y - 3, w, 3, fill=True, stroke=False)
    pdf.setFillColor(colors.HexColor("#667085"))
    pdf.setFont(regular_font, 8)
    pdf.drawString(x + 10, y - 18, title)
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 14)
    pdf.drawString(x + 10, y - 39, value)


def _draw_pdf_wrapped_text(pdf, text: str, x: int, y: int, max_width: int, regular_font: str, size: int = 9) -> int:
    pdf.setFillColor(colors.HexColor("#111827"))
    for line in wrap_text(pdf, text, max_width, regular_font, size):
        pdf.setFont(regular_font, size)
        pdf.drawString(x, y, line)
        y -= size + 5
    return y


def _draw_pdf_overview_table(pdf, x: int, y: int, rows: list[tuple[str, str]], width: int,
                              regular_font: str, bold_font: str) -> int:
    label_w = 100
    value_w = int(width - x * 2) - label_w
    row_h = 20
    for label, value in rows:
        pdf.setFillColor(colors.HexColor("#EAF4EF"))
        pdf.rect(x, y - row_h + 4, label_w, row_h, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#FFFFFF"))
        pdf.rect(x + label_w, y - row_h + 4, value_w, row_h, fill=True, stroke=False)
        pdf.setStrokeColor(colors.HexColor("#D9E2E8"))
        pdf.setLineWidth(0.4)
        pdf.rect(x, y - row_h + 4, label_w + value_w, row_h, fill=False, stroke=True)
        pdf.line(x + label_w, y - row_h + 4, x + label_w, y + 4)
        pdf.setFillColor(colors.HexColor("#1F2937"))
        pdf.setFont(bold_font, 9)
        pdf.drawString(x + 6, y - 9, label)
        pdf.setFont(regular_font, 9)
        pdf.drawString(x + label_w + 6, y - 9, str(value))
        y -= row_h
    return y - 8


def _draw_visual_pdf_cover(pdf, width: int, height: int, overview: dict, regular_font: str, bold_font: str):
    pdf.setFillColor(colors.white)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)
    draw_page_frame(pdf, width, height)

    pdf.setStrokeColor(colors.HexColor("#0F6B4F"))
    pdf.setLineWidth(1.2)
    pdf.line(75, height - 125, width - 75, height - 125)
    pdf.setLineWidth(0.6)
    pdf.setStrokeColor(colors.HexColor("#B8C7C1"))
    pdf.line(75, height - 132, width - 75, height - 132)

    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 21)
    pdf.drawCentredString(width / 2, height - 92, "AquaPaddy MRV 보고서")
    pdf.setFont(regular_font, 10)
    pdf.setFillColor(colors.HexColor("#344054"))
    pdf.drawCentredString(width / 2, height - 112, "IoT Water-Level Monitoring and Field Verification")

    logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png"))
    if os.path.exists(logo_path):
        try:
            pdf.drawImage(
                logo_path,
                width / 2 - 58,
                height - 320,
                width=116,
                height=116,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    report_month = overview.get("report_month") or "-"
    generated_at = overview.get("generated_at") or "-"
    generated_date = generated_at.split("T")[0] if isinstance(generated_at, str) else "-"
    field_name = overview.get("field_name") or "-"
    period = f"{overview.get('period_start', '-')} ~ {overview.get('period_end_exclusive', '-')}"

    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 16)
    pdf.drawCentredString(width / 2, height - 395, field_name)
    pdf.setFont(regular_font, 12)
    pdf.drawCentredString(width / 2, height - 422, f"{report_month} / MRV 기반 구축 보고서")

    box_x = 90
    box_y = height - 575
    box_w = width - 180
    box_h = 98
    pdf.setFillColor(colors.white)
    pdf.setStrokeColor(colors.HexColor("#D9E2E8"))
    pdf.rect(box_x, box_y, box_w, box_h, fill=True, stroke=True)
    pdf.setFont(regular_font, 10)
    pdf.setFillColor(colors.HexColor("#344054"))
    cover_rows = [
        ("연구 범위", "탄소배출권 거래 전 단계의 MRV 기반 구축"),
        ("분석 기간", period),
        ("작성일", generated_date),
        ("팀명", "강안장인"),
    ]
    row_y = box_y + box_h - 24
    for label, value in cover_rows:
        pdf.setFont(bold_font, 9)
        pdf.drawString(box_x + 22, row_y, label)
        pdf.setFont(regular_font, 9)
        pdf.drawString(box_x + 95, row_y, value)
        row_y -= 20

    pdf.setFillColor(colors.HexColor("#667085"))
    pdf.setFont(regular_font, 9)
    pdf.drawCentredString(width / 2, 142, "본 문서는 배출권 거래 결과물이 아닌 MRV 기반 구축 산출물입니다.")

    draw_page_number(pdf, width, 1, regular_font)


def _draw_visual_pdf_toc(pdf, width: int, height: int, regular_font: str, bold_font: str):
    draw_page_frame(pdf, width, height)
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 18)
    pdf.drawCentredString(width / 2, height - 72, "목차")
    pdf.setFont(regular_font, 9)
    pdf.setFillColor(colors.HexColor("#667085"))
    pdf.drawCentredString(width / 2, height - 94, "Table of Contents")

    toc_items = [
        ("1. 개요", "3", 0),
        ("2. 주요 내용 (결과 요약)", "3", 0),
        ("3. 결과 분석", "4", 0),
        ("3.1 주차별 수위 변화", "4", 1),
        ("3.2 현장 검증 결과", "4", 1),
        ("4. 결론", "5", 0),
    ]

    y = height - 155
    for title, page_no, indent in toc_items:
        x = LEFT_X + indent * 18
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont(bold_font if indent == 0 else regular_font, 11 if indent == 0 else 10)
        pdf.drawString(x, y, title)
        text_width = pdf.stringWidth(title, bold_font if indent == 0 else regular_font, 11 if indent == 0 else 10)
        page_width = pdf.stringWidth(page_no, regular_font, 10)
        dot_start = x + text_width + 10
        dot_end = width - RIGHT_MARGIN - page_width - 10
        pdf.setStrokeColor(colors.HexColor("#B8C7C1"))
        pdf.setDash(1, 3)
        pdf.line(dot_start, y + 3, dot_end, y + 3)
        pdf.setDash()
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont(regular_font, 10)
        pdf.drawRightString(width - RIGHT_MARGIN, y, page_no)
        y -= 30 if indent == 0 else 24

    pdf.setStrokeColor(colors.HexColor("#D9E2E8"))
    pdf.rect(LEFT_X, 120, int(width - LEFT_X - RIGHT_MARGIN), 54, fill=False, stroke=True)
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 10)
    pdf.drawString(LEFT_X + 16, 152, "보고서 작성 기준")
    pdf.setFont(regular_font, 9)
    pdf.setFillColor(colors.HexColor("#344054"))
    pdf.drawString(LEFT_X + 16, 134, "본 보고서는 배출권 거래 결과물이 아니라 MRV 기반 구축 산출물입니다.")

    draw_page_number(pdf, width, 2, regular_font)


def _draw_visual_mrv_pdf(report_id: int, db: Session):
    view = _mrv_view_data(report_id, db)
    overview = view["overview"]
    summary = view["summary"]
    validation = view["validation_results"]
    weekly_rows = view["weekly_analysis"]

    regular_font, bold_font = register_korean_fonts()
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    page_no = 1

    pdf.setTitle(f"AquaPaddy_MRV_Report_{report_id}")

    _draw_visual_pdf_cover(pdf, width, height, overview, regular_font, bold_font)
    pdf.showPage()
    page_no += 1

    _draw_visual_pdf_toc(pdf, width, height, regular_font, bold_font)
    pdf.showPage()
    page_no += 1

    y = _start_pdf_body_page(pdf, width, height)

    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 16)
    pdf.drawString(42, y, overview["field_name"])
    pdf.setFont(regular_font, 9)
    pdf.setFillColor(colors.HexColor("#667085"))
    pdf.drawString(42, y - 16,
                   f"분석기간: {overview['period_start']} ~ {overview['period_end_exclusive']}  |  보고월: {overview['report_month']}")
    pdf.setFillColor(colors.HexColor("#F7F9FB"))
    pdf.rect(42, y - 44, int(width - 84), 20, fill=True, stroke=False)
    pdf.setStrokeColor(colors.HexColor("#E5E7EB"))
    pdf.rect(42, y - 44, int(width - 84), 20, fill=False, stroke=True)
    pdf.setFillColor(colors.HexColor("#344054"))
    pdf.setFont(bold_font, 8)
    pdf.drawString(54, y - 38, "연구 포지션: 탄소배출권 거래 전 단계의 플랫폼 기반 구축")
    y -= 66

    card_w = 118
    gap = 9
    kpis = [
        ("AWD 사이클", _fmt_value(summary["total_awd_cycles"], "회")),
        ("Flood days", _fmt_value(summary["flood_days"], "일")),
        ("월 평균 수위", _fmt_value(summary["month_avg_inner_level_cm"], "cm")),
        ("탄소감축량", "미산정"),
    ]
    for i, (label, value) in enumerate(kpis):
        _draw_pdf_card(pdf, 42 + i * (card_w + gap), y, card_w, 56, label, value, regular_font, bold_font)
    y -= 76

    y = _draw_pdf_section_label(pdf, 42, y, "1. 개요", bold_font)
    loc = overview.get("field_location_desc") or "-"
    status_kor = "완료" if overview.get("status") == "COMPLETED" else "진행 중"
    overview_rows = [
        ("대상지", overview["field_name"]),
        ("위치", loc),
        ("분석기간", f"{overview['period_start']} ~ {overview['period_end_exclusive']}"),
        ("IoT 노드", f"{overview['node_count']}개"),
        ("보고 상태", status_kor),
    ]
    y = _draw_pdf_overview_table(pdf, 42, y, overview_rows, int(width), regular_font, bold_font)

    y = _draw_pdf_section_label(pdf, 42, y, "2. 주요 내용 (결과 요약)", bold_font)
    status_counts = summary["status_counts"]
    max_days = max(status_counts.values()) if status_counts else 1
    bar_total_w = 260
    for idx, status in enumerate(["OVERFLOODED", "FLOODED", "DRYING", "DRY"]):
        days = status_counts.get(status, 0)
        row_y = y - idx * 22
        pdf.setFillColor(colors.HexColor("#374151"))
        pdf.setFont(regular_font, 9)
        pdf.drawString(42, row_y - 1, _status_label(status))
        bar_w = int(bar_total_w * days / max_days) if max_days else 0
        pdf.setFillColor(colors.HexColor("#DDEFE8"))
        pdf.roundRect(42 + 72, row_y - 6, bar_total_w, 12, 4, fill=True, stroke=False)
        if bar_w > 0:
            pdf.setFillColor(colors.HexColor("#1D9E75"))
            pdf.roundRect(42 + 72, row_y - 6, bar_w, 12, 4, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont(bold_font, 9)
        pdf.drawString(42 + 72 + bar_total_w + 8, row_y - 2, f"{days}일")
    y -= 110

    _draw_pdf_footer(pdf, width, page_no, regular_font)
    pdf.showPage()
    page_no += 1

    y = _start_pdf_body_page(pdf, width, height)

    y = _draw_pdf_section_label(pdf, 42, y, "3. 결과 분석", bold_font)

    pdf.setFillColor(colors.HexColor("#444444"))
    pdf.setFont(bold_font, 10)
    pdf.drawString(42, y, "주차별 수위 변화")
    y -= 20

    w_headers = ["주차", "평균(cm)", "변화(cm)", "최소(cm)", "최대(cm)", "상태 흐름"]
    w_cols = [42, 62, 58, 58, 58, 212]
    row_h = 20
    table_x = 42
    table_w = sum(w_cols)

    pdf.setFillColor(colors.HexColor("#EAF4EF"))
    pdf.rect(table_x, y - row_h + 4, table_w, row_h, fill=True, stroke=False)
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont(bold_font, 8)
    cx = table_x
    for idx, h in enumerate(w_headers):
        pdf.drawString(cx + 4, y - 10, h)
        cx += w_cols[idx]
    pdf.setStrokeColor(colors.HexColor("#9BBCAE"))
    pdf.setLineWidth(0.6)
    pdf.rect(table_x, y - row_h + 4, table_w, row_h, fill=False, stroke=True)
    cx = table_x
    for col_w in w_cols[:-1]:
        cx += col_w
        pdf.line(cx, y + 4, cx, y - row_h + 4)
    y -= row_h

    pdf.setFont(regular_font, 8)
    for row_idx, row in enumerate(weekly_rows[:7]):
        flow_raw = row.get("status_flow") or "-"
        flow_safe = _flow_text(flow_raw).replace("→", "->")
        flow_lines = wrap_text(pdf, flow_safe, w_cols[5] - 8, regular_font, 8) or ["-"]
        current_row_h = max(row_h, 12 + len(flow_lines) * 10)
        if y - current_row_h < 70:
            _draw_pdf_footer(pdf, width, page_no, regular_font)
            pdf.showPage()
            page_no += 1
            y = _start_pdf_body_page(pdf, width, height)
        change = row.get("change_inner_level_cm")
        values = [
            f"{row['week_no']}주",
            _fmt_value(row["avg_inner_level_cm"]),
            f"{change:+.2f}" if change is not None else "-",
            _fmt_value(row["min_inner_level_cm"]),
            _fmt_value(row["max_inner_level_cm"]),
        ]
        if row_idx % 2 == 1:
            pdf.setFillColor(colors.HexColor("#F8FAFC"))
            pdf.rect(table_x, y - current_row_h + 4, table_w, current_row_h, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor("#111827"))
        cx = table_x
        for idx, val in enumerate(values):
            text = str(val)
            pdf.drawString(cx + 4, y - 10, text)
            cx += w_cols[idx]
        flow_x = table_x + sum(w_cols[:5]) + 4
        flow_y = y - 10
        for line in flow_lines:
            pdf.drawString(flow_x, flow_y, line)
            flow_y -= 10
        pdf.setStrokeColor(colors.HexColor("#C2CDC8"))
        pdf.setLineWidth(0.45)
        pdf.rect(table_x, y - current_row_h + 4, table_w, current_row_h, fill=False, stroke=True)
        cx = table_x
        for col_w in w_cols[:-1]:
            cx += col_w
            pdf.line(cx, y + 4, cx, y - current_row_h + 4)
        y -= current_row_h

    y -= 16

    pdf.setFillColor(colors.HexColor("#444444"))
    pdf.setFont(bold_font, 10)
    pdf.drawString(42, y, "현장 검증 결과")
    y -= 20

    v_kpis = [
        ("검증 샘플",      _fmt_value(validation["sample_count"], "건")),
        ("센서-관찰 일치율", _fmt_value(validation["sensor_observed_accuracy"], "%")),
        ("AI-센서 일치율",  _fmt_value(validation["ai_sensor_accuracy"], "%")),
        ("AI-센서 불일치",  _fmt_value(validation["ai_sensor_mismatch_count"], "건")),
    ]
    v_card_w = 118
    for i, (label, value) in enumerate(v_kpis):
        _draw_pdf_card(pdf, 42 + i * (v_card_w + gap), y, v_card_w, 52, label, value, regular_font, bold_font)
    y -= 68

    validation_note = validation.get("note") or ""
    if validation_note and validation_note != "별도 비고 없음":
        y = _draw_pdf_wrapped_text(pdf, f"메모: {validation_note}", 42, y, int(width - 84), regular_font, 9) - 10

    rep_rows = [r for r in validation.get("rows", []) if r.get("image_url")][:3]
    if rep_rows:
        pdf.setFont(regular_font, 8)
        pdf.setFillColor(colors.HexColor("#667085"))
        pdf.drawString(42, y, f"대표 검증 사진 ({len(rep_rows)}건)")
        y -= 14
        img_w = 150
        img_h = 112
        for i, row in enumerate(rep_rows):
            img_reader = load_image_reader(row["image_url"])
            img_x = 42 + i * (img_w + 8)
            if img_reader:
                try:
                    pdf.drawImage(img_reader, img_x, y - img_h, width=img_w, height=img_h,
                                  preserveAspectRatio=True, mask="auto")
                except Exception:
                    pdf.setFillColor(colors.HexColor("#F3F4F6"))
                    pdf.rect(img_x, y - img_h, img_w, img_h, fill=True, stroke=False)
            else:
                pdf.setFillColor(colors.HexColor("#F3F4F6"))
                pdf.rect(img_x, y - img_h, img_w, img_h, fill=True, stroke=False)
            pdf.setFont(regular_font, 7)
            pdf.setFillColor(colors.HexColor("#374151"))
            match_text = _match_label(row.get("ai_sensor_match"))
            pdf.drawString(img_x, y - img_h - 12, f"{row['record_date']}  AI-센서: {match_text}")
        y -= img_h + 22

    _draw_pdf_footer(pdf, width, page_no, regular_font)
    pdf.showPage()
    page_no += 1

    y = _start_pdf_body_page(pdf, width, height)

    y = _draw_pdf_section_label(pdf, 42, y, "4. 결론", bold_font)

    conclusion_lines = view.get("conclusion", [])
    wrapped_line_count = sum(
        max(1, len(wrap_text(pdf, line, int(width - 112), regular_font, 10)))
        for line in conclusion_lines
    )
    box_h = 20 + wrapped_line_count * 15 + max(0, len(conclusion_lines) - 1) * 4
    pdf.setFillColor(colors.HexColor("#EAF4EF"))
    pdf.setStrokeColor(colors.HexColor("#B2DAC9"))
    pdf.setLineWidth(0.6)
    pdf.roundRect(42, y - box_h, int(width - 84), box_h, 6, fill=True, stroke=True)
    cy = y - 14
    for line in conclusion_lines:
        y_after = _draw_pdf_wrapped_text(pdf, line, 56, cy, int(width - 112), regular_font, 10)
        cy = y_after - 2
    y = y - box_h - 18

    summary_table_rows = [
        ("AWD 사이클", _fmt_value(summary["total_awd_cycles"], "회")),
        ("Flood days", _fmt_value(summary["flood_days"], "일")),
        ("월 평균 수위", _fmt_value(summary["month_avg_inner_level_cm"], "cm")),
        ("탄소감축량", "미산정"),
        ("검증 샘플", _fmt_value(validation["sample_count"], "건")),
        ("AI-센서 일치율", _fmt_value(validation["ai_sensor_accuracy"], "%")),
    ]
    y = _draw_pdf_wrapped_text(pdf, "주요 수치 요약", 42, y, int(width - 84), bold_font, 10) - 8
    y = _draw_pdf_overview_table(pdf, 42, y, summary_table_rows, int(width), regular_font, bold_font)

    _draw_pdf_footer(pdf, width, page_no, regular_font)
    pdf.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="mrv_report_{report_id}.pdf"'},
    )


def _style_excel_cells(sheet, cell_range: str, fill=None, font=None, alignment=None, border=None):
    for row in sheet[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border


def _draw_visual_mrv_excel(report_id: int, db: Session):
    view = _mrv_view_data(report_id, db)
    overview = view["overview"]
    summary = view["summary"]
    validation = view["validation_results"]

    workbook = Workbook()
    workbook.remove(workbook.active)

    title_font = XLFont(name="Malgun Gothic", size=18, bold=True, color="FFFFFF")
    subtitle_font = XLFont(name="Malgun Gothic", size=11, bold=True, color="0F6B4F")
    section_font = XLFont(name="Malgun Gothic", size=12, bold=True, color="111827")
    header_font = XLFont(name="Malgun Gothic", size=10, bold=True, color="111827")
    body_font = XLFont(name="Malgun Gothic", size=10, color="111827")
    muted_font = XLFont(name="Malgun Gothic", size=9, color="667085")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    green_fill = PatternFill("solid", fgColor="0F6B4F")
    light_green_fill = PatternFill("solid", fgColor="EAF4EF")
    light_blue_fill = PatternFill("solid", fgColor="EAF3FF")
    gray_fill = PatternFill("solid", fgColor="F7F9FB")
    border = Border(
        left=Side(style="thin", color="D9E2E8"),
        right=Side(style="thin", color="D9E2E8"),
        top=Side(style="thin", color="D9E2E8"),
        bottom=Side(style="thin", color="D9E2E8"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    report_sheet = workbook.create_sheet("MRV Report")
    report_sheet.sheet_view.showGridLines = False
    for col, width in {"A": 4, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18}.items():
        report_sheet.column_dimensions[col].width = width

    report_sheet.merge_cells("A1:H2")
    report_sheet["A1"] = "MRV 보고서"
    report_sheet["A1"].font = title_font
    report_sheet["A1"].fill = green_fill
    report_sheet["A1"].alignment = left
    _style_excel_cells(report_sheet, "A1:H2", fill=green_fill, font=title_font, alignment=left)

    report_sheet.merge_cells("A3:H3")
    report_sheet["A3"] = f"{overview['field_name']} · {overview['period_start']} ~ {overview['period_end_exclusive']}"
    report_sheet["A3"].font = subtitle_font
    report_sheet["A3"].alignment = left

    kpis = [
        ("AWD 사이클", _fmt_value(summary["total_awd_cycles"], "회")),
        ("Flood days", _fmt_value(summary["flood_days"], "일")),
        ("월 평균 수위", _fmt_value(summary["month_avg_inner_level_cm"], "cm")),
        ("탄소감축량", "미산정"),
    ]
    row = 5
    for idx, (label, value) in enumerate(kpis):
        start_col = 1 + idx * 2
        end_col = start_col + 1
        report_sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        report_sheet.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 2, end_column=end_col)
        label_cell = report_sheet.cell(row=row, column=start_col, value=label)
        value_cell = report_sheet.cell(row=row + 1, column=start_col, value=value)
        _style_excel_cells(
            report_sheet,
            f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row + 2}",
            fill=gray_fill,
            border=border,
            alignment=center,
        )
        label_cell.font = muted_font
        value_cell.font = XLFont(name="Malgun Gothic", size=16, bold=True, color="111827")
        value_cell.alignment = center

    row = 9
    report_sheet.merge_cells(f"A{row}:H{row}")
    report_sheet[f"A{row}"] = "1. 개요"
    _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=light_green_fill, font=section_font, alignment=left, border=border)
    row += 1
    overview_rows = [
        ("대상지", overview["field_name"]),
        ("위치", overview.get("field_location_desc") or "-"),
        ("분석기간", f"{overview['period_start']} ~ {overview['period_end_exclusive']}"),
        ("노드 수", _fmt_value(overview["node_count"], "개")),
        ("보고 상태", "완료" if overview["status"] == "COMPLETED" else "진행 중"),
    ]
    for label, value in overview_rows:
        report_sheet.merge_cells(f"A{row}:B{row}")
        report_sheet.merge_cells(f"C{row}:H{row}")
        report_sheet[f"A{row}"] = label
        report_sheet[f"C{row}"] = value
        _style_excel_cells(report_sheet, f"A{row}:B{row}", fill=light_blue_fill, font=header_font, alignment=center, border=border)
        _style_excel_cells(report_sheet, f"C{row}:H{row}", fill=white_fill, font=body_font, alignment=left, border=border)
        row += 1

    row += 1
    report_sheet.merge_cells(f"A{row}:H{row}")
    report_sheet[f"A{row}"] = "2. 주요 내용"
    _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=light_green_fill, font=section_font, alignment=left, border=border)
    row += 1
    report_sheet.append(["상태", "일수", "", "", "검증 항목", "값", "", ""])
    _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=light_blue_fill, font=header_font, alignment=center, border=border)
    row += 1
    status_counts = summary["status_counts"]
    validation_metrics = [
        ("검증 샘플", _fmt_value(validation["sample_count"], "건")),
        ("센서-관찰 일치율", _fmt_value(validation["sensor_observed_accuracy"], "%")),
        ("AI-센서 일치율", _fmt_value(validation["ai_sensor_accuracy"], "%")),
        ("AI-센서 불일치", _fmt_value(validation["ai_sensor_mismatch_count"], "건")),
    ]
    for idx, status in enumerate(["OVERFLOODED", "FLOODED", "DRYING", "DRY"]):
        metric_label, metric_value = validation_metrics[idx]
        values = [_status_label(status), status_counts.get(status, 0), "", "", metric_label, metric_value, "", ""]
        for col_idx, value in enumerate(values, start=1):
            cell = report_sheet.cell(row=row, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = center if col_idx in (1, 2, 5, 6) else left
            cell.border = border
        row += 1

    row += 1
    report_sheet.merge_cells(f"A{row}:H{row}")
    report_sheet[f"A{row}"] = "3. 결과 분석"
    _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=light_green_fill, font=section_font, alignment=left, border=border)
    row += 1
    weekly_main_headers = ["주차", "평균 수위(cm)", "변화(cm)", "최소(cm)", "최대(cm)", "상태 흐름", "", ""]
    for col_idx, value in enumerate(weekly_main_headers, start=1):
        cell = report_sheet.cell(row=row, column=col_idx, value=value)
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.alignment = center
        cell.border = border
    row += 1
    if view["weekly_analysis"]:
        for row_data in view["weekly_analysis"]:
            change = row_data.get("change_inner_level_cm")
            values = [
                f"{row_data['week_no']}주",
                row_data["avg_inner_level_cm"],
                f"{change:+.2f}" if change is not None else "-",
                row_data["min_inner_level_cm"],
                row_data["max_inner_level_cm"],
                _flow_text(row_data["status_flow"]),
                "",
                "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = report_sheet.cell(row=row, column=col_idx, value=value)
                cell.font = body_font
                cell.fill = white_fill
                cell.alignment = left if col_idx == 6 else center
                cell.border = border
            row += 1
    else:
        report_sheet.merge_cells(f"A{row}:H{row}")
        report_sheet[f"A{row}"] = "주차별 수위 데이터 없음"
        _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=white_fill, font=body_font, alignment=center, border=border)
        row += 1

    report_sheet.merge_cells(f"A{row}:H{row}")
    report_sheet[f"A{row}"] = (
        f"현장 검증 결과: 검증 샘플 {_fmt_value(validation['sample_count'], '건')}, "
        f"센서-관찰 일치율 {_fmt_value(validation['sensor_observed_accuracy'], '%')}, "
        f"AI-센서 일치율 {_fmt_value(validation['ai_sensor_accuracy'], '%')}"
    )
    _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=gray_fill, font=body_font, alignment=left, border=border)

    row += 2
    report_sheet.merge_cells(f"A{row}:H{row}")
    report_sheet[f"A{row}"] = "4. 결론"
    _style_excel_cells(report_sheet, f"A{row}:H{row}", fill=light_green_fill, font=section_font, alignment=left, border=border)
    row += 1
    report_sheet.merge_cells(f"A{row}:H{row + 3}")
    report_sheet[f"A{row}"] = "\n".join(view["conclusion"])
    _style_excel_cells(report_sheet, f"A{row}:H{row + 3}", fill=white_fill, font=body_font, alignment=left, border=border)
    report_sheet.row_dimensions[row].height = 78

    weekly_sheet = workbook.create_sheet("Weekly Analysis")
    weekly_sheet.sheet_view.showGridLines = False
    weekly_headers = ["주차", "평균 수위(cm)", "변화(cm)", "최소 수위(cm)", "최대 수위(cm)", "상태 흐름"]
    weekly_sheet.append(weekly_headers)
    _style_excel_cells(weekly_sheet, "A1:F1", fill=green_fill, font=XLFont(name="Malgun Gothic", size=10, bold=True, color="FFFFFF"), alignment=center, border=border)
    for row_data in view["weekly_analysis"]:
        change = row_data.get("change_inner_level_cm")
        weekly_sheet.append([
            f"{row_data['week_no']}주",
            row_data["avg_inner_level_cm"],
            f"{change:+.2f}" if change is not None else "-",
            row_data["min_inner_level_cm"],
            row_data["max_inner_level_cm"],
            _flow_text(row_data["status_flow"]),
        ])
    for row_cells in weekly_sheet.iter_rows(min_row=2, max_row=weekly_sheet.max_row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.font = body_font
            cell.alignment = left if cell.column == 6 else center
            cell.border = border
    for col, width in {"A": 10, "B": 16, "C": 14, "D": 16, "E": 16, "F": 70}.items():
        weekly_sheet.column_dimensions[col].width = width

    validation_sheet = workbook.create_sheet("Validation")
    validation_sheet.sheet_view.showGridLines = False
    validation_headers = ["일자", "노드", "관찰 상태", "센서 상태", "AI 상태", "AI 신뢰도", "센서-관찰", "AI-센서", "이미지 URL", "메모"]
    validation_sheet.append(validation_headers)
    _style_excel_cells(validation_sheet, "A1:J1", fill=green_fill, font=XLFont(name="Malgun Gothic", size=10, bold=True, color="FFFFFF"), alignment=center, border=border)
    if validation["rows"]:
        for row_data in validation["rows"]:
            validation_sheet.append([
                row_data["record_date"],
                row_data["node_id"],
                row_data["observed_surface_status"],
                row_data["sensor_predicted_status"],
                row_data["ai_predicted_status"],
                row_data["ai_confidence"],
                _match_label(row_data["sensor_observed_match"]),
                _match_label(row_data["ai_sensor_match"]),
                row_data["image_url"],
                row_data["note"],
            ])
    else:
        validation_sheet.append(["검증 데이터 없음", "", "", "", "", "", "", "", "", ""])
    for row_cells in validation_sheet.iter_rows(min_row=2, max_row=validation_sheet.max_row, min_col=1, max_col=10):
        for cell in row_cells:
            cell.font = body_font
            cell.alignment = left if cell.column in (9, 10) else center
            cell.border = border
    for col, width in {"A": 14, "B": 10, "C": 16, "D": 16, "E": 14, "F": 12, "G": 14, "H": 14, "I": 50, "J": 32}.items():
        validation_sheet.column_dimensions[col].width = width

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="mrv_report_{report_id}_visual.xlsx"'},
    )


@router.get("/{report_id}/download/pdf")
def download_mrv_report_pdf(report_id: int, db: Session = Depends(get_db)):
    return _draw_visual_mrv_pdf(report_id, db)

    report = db.query(MrvReport).filter(MrvReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="해당 report_id가 존재하지 않습니다.")

    field = db.query(Field).filter(Field.id == report.field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="해당 보고서의 논 정보가 존재하지 않습니다.")

    start_date, end_date = get_month_range(report.report_month)
    validation = get_validation_summary(report.field_id, start_date, end_date, db)

    summaries = (
        db.query(AwdDailySummary)
        .join(IotNode, IotNode.id == AwdDailySummary.node_id)
        .filter(
            IotNode.field_id == report.field_id,
            AwdDailySummary.record_date >= start_date,
            AwdDailySummary.record_date < end_date,
        )
        .order_by(AwdDailySummary.record_date.asc(), AwdDailySummary.id.asc())
        .all()
    )

    nodes = db.query(IotNode).filter(IotNode.field_id == report.field_id).order_by(IotNode.id.asc()).all()
    daily_summaries = aggregate_daily_summaries(summaries)
    status_counts = summarize_status_counts(daily_summaries)
    weekly_groups = group_summaries_by_week(daily_summaries)
    weekly_dict = {week_no: week_summaries for week_no, week_summaries in weekly_groups}
    month_avg = get_month_avg_inner_level(daily_summaries)
    dominant_status = get_dominant_status(status_counts)
    report_month_kor = format_report_month(report.report_month)

    regular_font, bold_font = register_korean_fonts()

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    max_text_width = int(width - LEFT_X - RIGHT_MARGIN)
    pdf.setTitle(f"mrv_report_{report.id}")
    page_no_ref = [1]

    draw_page_frame(pdf, width, height)

    pdf.setFont(bold_font, COVER_TITLE_SIZE)
    pdf.drawCentredString(width / 2, height - 110, "AWD Water Management MRV Report")

    pdf.setLineWidth(1.0)
    pdf.setStrokeColor(colors.black)
    pdf.line(75, height - 135, width - 75, height - 135)

    logo_path = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")
    )

    if os.path.exists(logo_path):
        pdf.drawImage(
            logo_path,
            width / 2 - 80,
            height - 330,
            width=160,
            height=160,
            preserveAspectRatio=True,
            mask="auto"
        )

    pdf.setFont(bold_font, COVER_SUBTITLE_SIZE)
    pdf.drawCentredString(width / 2, height - 420, field.field_name)
    pdf.drawCentredString(width / 2, height - 455, f"{report_month_kor} MRV 보고서")

    pdf.setFont(regular_font, COVER_INFO_SIZE)
    created_text = str(report.created_at.date()) if report.created_at else "-"
    pdf.drawCentredString(width / 2, height - 545, f"작성일: {created_text}")
    pdf.drawCentredString(width / 2, height - 570, "팀명: 강안장인")

    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.showPage()
    page_no_ref[0] += 1

    pdf.setFont(bold_font, TOC_TITLE_SIZE)
    pdf.drawCentredString(width / 2, height - 55, "목차")

    toc_items = [
        ("1. 개요", "3", 0),
        ("2. 주요 내용 (결과 요약)", "3", 0),
        ("3. 결과 분석", "4", 0),
        ("3.1 주차별 수위 변화", "4", 1),
        ("3.2 현장 검증 결과", "4", 1),
        ("4. 결론", "5", 0),
    ]

    y = height - 125
    pdf.setFont(regular_font, TOC_BODY_SIZE)
    for title, page_no, indent in toc_items:
        x = LEFT_X + indent * 18
        pdf.drawString(x, y, title)
        text_width = pdf.stringWidth(title, regular_font, TOC_BODY_SIZE)
        page_width = pdf.stringWidth(page_no, regular_font, TOC_BODY_SIZE)
        dot_start = x + text_width + 8
        dot_end = width - RIGHT_MARGIN - page_width - 8
        pdf.setDash(1, 2)
        pdf.line(dot_start, y + 2, dot_end, y + 2)
        pdf.setDash()
        pdf.drawRightString(width - RIGHT_MARGIN, y, page_no)
        y -= 25
    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.showPage()
    page_no_ref[0] += 1

    y = TOP_Y

    y = draw_section_title(pdf, "1. 개요", y, regular_font, bold_font)
    y = draw_text(
        pdf,
        f"본 보고서는 {field.field_name}을 대상으로 {report_month_kor} 동안 수행된 "
        "AWD(Alternate Wetting and Drying) 물관리 이력을 IoT 센서 기반으로 분석한 MRV 보고서이다.",
        LEFT_X, y, max_text_width, regular_font, BODY_SIZE, BODY_LINE_HEIGHT,
    )
    y -= 4
    y = draw_simple_table(
        pdf, LEFT_X, y,
        ["항목", "내용"],
        [
            ["대상지", field.field_name],
            ["위치", field.location_desc or "-"],
            ["분석 기간", report_month_kor],
            ["사용 노드 수", f"{len(nodes)}개"],
            ["보고서 생성일", created_text],
        ],
        [150, 300], regular_font, bold_font,
    )
    y = draw_text(pdf, "수위 상태는 내부 수위를 기준으로 다음과 같이 구분된다.", LEFT_X, y, max_text_width, regular_font)
    y = draw_bullets(
        pdf,
        [
            "OVERFLOODED",
            "FLOODED",
            "DRYING",
            "DRY: 재관개 필요 상태",
        ],
        LEFT_X + 8, y, max_text_width - 8, regular_font,
    )

    y = draw_divider(pdf, y - 4, width)

    y = draw_section_title(pdf, "2. 주요 내용 (결과 요약)", y, regular_font, bold_font)
    y = draw_simple_table(
        pdf, LEFT_X, y,
        ["AWD 사이클", "Flood days", "월 평균 수위", "탄소 감축량"],
        [[
            f"{report.total_awd_cycles}회",
            f"{report.flood_days}일",
            f"{month_avg:.2f}cm" if month_avg is not None else "-",
            f"{report.carbon_reduction} kgCO2-eq",
        ]],
        [110, 100, 130, 150], regular_font, bold_font,
    )
    y -= 4
    y = draw_sub_title(pdf, "상태 분포 (일수)", y, bold_font)
    y = draw_status_bar_chart(pdf, LEFT_X, y, status_counts, regular_font, bold_font)
    y -= 6
    y = draw_text(
        pdf,
        f"분석 기간 동안 {dominant_status} 상태가 가장 많이 관측되었으며, AWD 수행 횟수는 {report.total_awd_cycles}회로 나타났다. "
        f"탄소 감축량은 AWD 수행 횟수 × 15.25(kgCO2-eq) 기준으로 {report.carbon_reduction} kgCO2-eq로 산정되었다. "
        "(센서 기반 수위 데이터 분석에 따른 추정값)",
        LEFT_X, y, max_text_width, regular_font,
    )
    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.showPage()
    page_no_ref[0] += 1

    y = TOP_Y
    y = draw_section_title(pdf, "3. 결과 분석", y, regular_font, bold_font)

    y = draw_sub_title(pdf, "3.1 주차별 수위 변화", y, bold_font)
    weekly_rows = build_weekly_table_rows(weekly_dict)
    if weekly_rows:
        y = draw_simple_table(
            pdf, LEFT_X, y,
            ["주차", "평균", "변화", "최소", "최대", "상태 흐름"],
            weekly_rows,
            [40, 60, 60, 55, 55, 220], regular_font, bold_font,
        )
        weekly_avgs = [float(r[1]) for r in weekly_rows]
    else:
        monthly_text = "분석 기간 동안 평균 내부 수위 데이터가 없어 월간 수위 상태 분석이 제한된다."
    y = draw_text(pdf, monthly_text, LEFT_X, y, max_text_width, regular_font, BODY_SIZE, BODY_LINE_HEIGHT)
    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.showPage()
    page_no_ref[0] += 1

    # Page 3: 5
    y = TOP_Y
    y = draw_section_title(pdf, "5. AWD 수행 분석 및 탄소감축 산정 기반", y, regular_font, bold_font)

    y = draw_sub_title(pdf, "[AWD 수행 횟수 기준]", y, bold_font)
    y = draw_text(pdf, "논이 DRY 상태 이후 DRYING, FLOODED 또는 OVERFLOODED 상태로 전환되는 경우를 1회로 정의한다.", LEFT_X, y, max_text_width, regular_font)
    y = draw_bullets(pdf, [f"{report_month_kor} AWD 수행 횟수: {report.total_awd_cycles}회"], LEFT_X + 8, y - 4, max_text_width, regular_font)
    y -= 14

    y = draw_sub_title(pdf, "[탄소감축량 산정 범위]", y, bold_font)
    y = draw_text(
        pdf,
        "본 연구에서는 실제 메탄 배출량 측정 및 공식 탄소감축량 산정까지는 수행하지 않았다.\n"
        "다만 AWD 수행 횟수, 수위 변화, 검증 결과를 기록함으로써 향후 탄소감축량 산정을 위한 기초 데이터를 확보하였다.",
        LEFT_X, y, max_text_width, regular_font,
    )
    y -= 14

    y = draw_sub_title(pdf, "[시사점]", y, bold_font)
    if report.total_awd_cycles == 0:
        insight_text = (
            "건조 단계 이후 재관개가 이루어지지 않아 AWD 사이클이 형성되지 않았다.\n"
            "이는 물관리 전략이 건조 단계 중심으로 운영되었음을 의미한다.\n"
            "향후 AWD 수행을 위해서는 DRY 상태 이후 적절한 시점에서의 계획적 재관개가 필요하다."
        )
    else:
        insight_text = (
            f"분석 기간 동안 AWD 사이클이 {report.total_awd_cycles}회 확인되었다.\n"
            "이는 건조 이후 재관개가 수행되어 AWD 물관리 흐름이 일부 형성되었음을 의미한다.\n"
            "향후에는 주기적인 검증 데이터 확보를 통해 AWD 수행 결과의 신뢰성을 높일 필요가 있다."
        )
    y = draw_text(pdf, insight_text, LEFT_X, y, max_text_width, regular_font)
    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.showPage()
    page_no_ref[0] += 1

    # Page 4+: 6 검증 결과. 사진이 길면 자동 페이지 넘김.
    y = TOP_Y
    y = draw_section_title(pdf, "6. 검증 결과", y, regular_font, bold_font)
    y = draw_sub_title(pdf, "6.1 현장 검증 결과", y, bold_font)

    if validation["validation_sample_count"] > 0:
        y = draw_text(
            pdf,
            f"주차별 평균 내부 수위는 {min(weekly_avgs):.2f}cm ~ {max(weekly_avgs):.2f}cm 범위에서 변동하였으며, "
            f"분석 기간 전반에 걸쳐 {dominant_status} 상태가 우세하게 나타났다.",
            LEFT_X, y, max_text_width, regular_font,
        )
    else:
        y = draw_text(pdf, "주차별 수위 데이터가 부족하여 상세 분석이 제한된다.", LEFT_X, y, max_text_width, regular_font)

    y -= 12

    y = draw_sub_title(pdf, "3.2 현장 검증 결과", y, bold_font)
    if validation["validation_sample_count"] > 0:
        y = draw_simple_table(
            pdf, LEFT_X, y,
            ["항목", "값"],
            [
                ["검증 방법", validation["validation_method"]],
                ["검증 샘플 수", f"{validation['validation_sample_count']}건"],
                ["센서-관찰 일치 / 불일치", f"{validation['validation_match_count']}건 / {validation['validation_mismatch_count']}건"],
                ["센서-관찰 정확도", f"{validation['validation_accuracy']}%"],
                ["AI-센서 일치 / 불일치", f"{validation['ai_sensor_match_count']}건 / {validation['ai_sensor_mismatch_count']}건"],
                ["AI-센서 일치율", f"{validation['ai_sensor_accuracy']}%"],
            ],
            [180, 240], regular_font, bold_font,
        )
        y = draw_text(
            pdf,
            "검증 결과는 센서 기반 상태 판정의 신뢰성을 확인하기 위한 보조 자료로 활용된다. "
            "AI-센서 일치율은 촬영 시점 근접 센서 로그의 수위값을 LOW/MID/HIGH 구간으로 변환한 뒤 AI 예측 결과와 비교하여 산정하였다.",
            LEFT_X, y, max_text_width, regular_font,
        )
    else:
        y = draw_text(
            pdf,
            "해당 기간 동안 등록된 현장 검증 데이터가 없어 검증 결과 분석이 제한된다.",
            LEFT_X, y, max_text_width, regular_font,
        )

    representative_rows = validation["representative_rows"]
    if representative_rows:
        y -= 16
        y = ensure_space_for_validation(pdf, y, 340, width, height, regular_font, page_no_ref)
        y = draw_sub_title(pdf, "대표 검증 이미지 (월 초·월 중·월 말)", y, bold_font)
        card_width = 160
        card_height = 245
        gap = 12
        card_y = y

        for idx, (label, row) in enumerate(representative_rows[:3]):
            x = LEFT_X + idx * (card_width + gap)

            pdf.setStrokeColor(colors.lightgrey)
            pdf.setLineWidth(0.5)
            pdf.roundRect(x, card_y - card_height, card_width, card_height, 8, stroke=True, fill=False)

            pdf.setFillColor(colors.black)
            pdf.setFont(bold_font, 10)
            pdf.drawString(x + 8, card_y - 18, f"{label} 대표 이미지")

            img = load_image_reader(row.image_url)
            if img:
                pdf.drawImage(
                    img, x + 8, card_y - 120, width=card_width - 16, height=90,
                    preserveAspectRatio=True, mask="auto",
                )
            else:
                pdf.setFillColor(colors.HexColor("#F3F4F6"))
                pdf.roundRect(x + 8, card_y - 120, card_width - 16, 90, 4, fill=True, stroke=False)
                pdf.setFillColor(colors.HexColor("#9CA3AF"))
                pdf.setFont(regular_font, 8)
                pdf.drawCentredString(x + card_width / 2, card_y - 80, "이미지 없음")

            sensor_ai_status = get_sensor_ai_status_for_validation(row, db) or "-"
            ai_status = row.ai_predicted_status or "-"
            observed = row.observed_surface_status or "-"
            sensor_surface = row.sensor_predicted_status or "-"
            match_text = validation_result_text(row.is_match)
            ai_match_text = validation_result_text(row.ai_sensor_match)

            info_groups = [
                [f"촬영일: {row.record_date}"],
                [f"사람 관찰: {observed}", f"센서 표면: {sensor_surface}"],
                [f"센서 구간: {sensor_ai_status}", f"AI 구간: {ai_status}"],
                [f"센서-사람: {match_text}", f"AI-센서: {ai_match_text}"],
            ]

            text_y = card_y - 138
            pdf.setFillColor(colors.black)
            pdf.setFont(regular_font, 8)
            for group_idx, group in enumerate(info_groups):
                for line in group:
                    pdf.drawString(x + 8, text_y, line)
                    text_y -= 11
                if group_idx < len(info_groups) - 1:
                    pdf.setStrokeColor(colors.lightgrey)
                    pdf.setLineWidth(0.3)
                    pdf.line(x + 8, text_y + 3, x + card_width - 8, text_y + 3)
                    text_y -= 6

        y = card_y - card_height - 20

    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.showPage()
    page_no_ref[0] += 1
    y = TOP_Y
    y = draw_section_title(pdf, "4. 결론", y, regular_font, bold_font)
    if report.total_awd_cycles == 0:
        conclusion = (
            f"본 분석 기간 동안 논은 전반적으로 {dominant_status} 상태를 중심으로 변화하였으며, AWD 수행은 발생하지 않았다. "
            "이는 건조 이후 재관개가 이루어지지 않았기 때문으로 판단된다.\n"
            "본 시스템을 통해 IoT 기반 수위 데이터 수집, 상태 분석, 현장 검증, MRV 보고서 생성까지의 자동화 가능성을 확인하였다. "
            "본 보고서는 탄소배출권 거래 또는 공식 감축량 산정을 완료한 결과물이 아니라, "
            "향후 탄소감축량 산정 및 탄소배출권 제도 연계를 위한 MRV 기반 자료로 활용될 수 있다."
        )
    else:
        conclusion = (
            f"본 분석 기간 동안 AWD 수행은 {report.total_awd_cycles}회 관측되었으며, "
            "수위 데이터 수집·상태 분석·현장 검증·보고서 생성 과정을 자동화하였다.\n"
            "본 보고서는 탄소배출권 거래 또는 공식 감축량 산정을 완료한 결과물이 아니라, "
            "향후 탄소감축량 산정 및 탄소배출권 제도 연계를 위한 MRV 기반 자료로 활용될 수 있다."
        )
    y = draw_text(pdf, conclusion, LEFT_X, y, max_text_width, regular_font)

    y -= 14
    y = draw_sub_title(pdf, "[향후 계획]", y, bold_font)
    y = draw_bullets(
        pdf,
        [
            "AWD 사이클 확보를 위한 계획적 재관개 전략 수립",
            "IoT 센서 기반 자동 관개 시스템 도입 검토",
            "현장 검증 데이터(이미지) 수집 및 검증 체계 강화",
            "MRV 보고서 자동화 및 시각화 기능 개선",
        ],
        LEFT_X + 8, y, max_text_width, regular_font,
    )

    draw_page_number(pdf, width, page_no_ref[0], regular_font)
    pdf.save()
    buffer.seek(0)

    filename = f"mrv_report_{report.id}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{report_id}/download/excel")
def download_mrv_report_excel(report_id: int, db: Session = Depends(get_db)):
    return _draw_visual_mrv_excel(report_id, db)

    report = db.query(MrvReport).filter(MrvReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="해당 report_id가 존재하지 않습니다.")

    field = db.query(Field).filter(Field.id == report.field_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="해당 보고서의 논 정보가 존재하지 않습니다.")

    start_date, end_date = get_month_range(report.report_month)
    validation = get_validation_summary(report.field_id, start_date, end_date, db)
    validation_rows = get_validation_rows(report.field_id, start_date, end_date, db)

    summaries = (
        db.query(AwdDailySummary)
        .join(IotNode, IotNode.id == AwdDailySummary.node_id)
        .filter(
            IotNode.field_id == report.field_id,
            AwdDailySummary.record_date >= start_date,
            AwdDailySummary.record_date < end_date,
        )
        .order_by(AwdDailySummary.record_date.asc(), AwdDailySummary.id.asc())
        .all()
    )

    nodes = db.query(IotNode).filter(IotNode.field_id == report.field_id).order_by(IotNode.id.asc()).all()
    daily_summaries = aggregate_daily_summaries(summaries)
    status_counts = summarize_status_counts(daily_summaries)
    dominant_status = get_dominant_status(status_counts)
    month_avg = get_month_avg_inner_level(daily_summaries)
    report_month_kor = format_report_month(report.report_month)
    created_text = str(report.created_at.date()) if report.created_at else "-"

    workbook = Workbook()

    title_font = XLFont(name="맑은 고딕", size=16, bold=True)
    section_font = XLFont(name="맑은 고딕", size=12, bold=True)
    header_font = XLFont(name="맑은 고딕", size=10, bold=True)
    body_font = XLFont(name="맑은 고딕", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="808080")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    section_fill = PatternFill("solid", fgColor="EAF1F8")
    note_fill = PatternFill("solid", fgColor="F7F7F7")

    def set_widths(sheet, widths: dict[str, float]):
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width

    def style_range(sheet, cell_range: str, fill=None, font=None, alignment=None, border_style=True):
        for row in sheet[cell_range]:
            for cell in row:
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if border_style:
                    cell.border = border

    def put_section_title(sheet, row: int, title: str, last_col: str = "H") -> int:
        sheet.merge_cells(f"A{row}:{last_col}{row}")
        cell = sheet[f"A{row}"]
        cell.value = title
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = left
        style_range(sheet, f"A{row}:{last_col}{row}", fill=section_fill, font=section_font, alignment=left)
        sheet.row_dimensions[row].height = 24
        return row + 1

    def put_kv(sheet, row: int, key: str, value, key_range: str, value_range: str) -> int:
        sheet.merge_cells(key_range.format(row=row))
        sheet.merge_cells(value_range.format(row=row))
        key_cell = sheet[f"A{row}"]
        value_col = value_range.split("{")[0].split(":")[-1].rstrip("0123456789") or "C"
        value_cell_ref = value_range.format(row=row).split(":")[0]
        value_cell = sheet[value_cell_ref]
        key_cell.value = key
        value_cell.value = value
        style_range(sheet, key_range.format(row=row), fill=header_fill, font=header_font, alignment=center)
        style_range(sheet, value_range.format(row=row), font=body_font, alignment=left)
        sheet.row_dimensions[row].height = 22
        return row + 1

    summary_sheet = workbook.active
    summary_sheet.title = "요약"
    summary_sheet.sheet_view.showGridLines = False
    set_widths(summary_sheet, {
        "A": 4, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 18
    })

    summary_sheet.merge_cells("A1:H1")
    summary_sheet["A1"] = "AWD Water Management MRV Report"
    summary_sheet["A1"].font = title_font
    summary_sheet["A1"].alignment = center
    summary_sheet.row_dimensions[1].height = 32

    summary_sheet.merge_cells("A2:H2")
    summary_sheet["A2"] = f"{field.field_name} / {report_month_kor} MRV 보고서"
    summary_sheet["A2"].font = XLFont(name="맑은 고딕", size=11, bold=True)
    summary_sheet["A2"].alignment = center
    summary_sheet.row_dimensions[2].height = 24

    row = 4
    row = put_section_title(summary_sheet, row, "1. 분석 대상 및 기간")
    row = put_kv(summary_sheet, row, "대상 논", field.field_name, "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "분석 기간", report_month_kor, "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "사용 노드 수", f"{len(nodes)}개", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "보고서 생성일", created_text, "A{row}:B{row}", "C{row}:H{row}")

    row += 1
    row = put_section_title(summary_sheet, row, "2. 결과 요약")
    summary_sheet.merge_cells(f"A{row}:D{row}")
    summary_sheet.merge_cells(f"E{row}:H{row}")
    summary_sheet[f"A{row}"] = "상태"
    summary_sheet[f"E{row}"] = "일수"
    style_range(summary_sheet, f"A{row}:H{row}", fill=header_fill, font=header_font, alignment=center)
    row += 1
    for status, days in [
        ("OVERFLOODED", status_counts["OVERFLOODED"]),
        ("FLOODED", status_counts["FLOODED"]),
        ("DRYING", status_counts["DRYING"]),
        ("DRY", status_counts["DRY"]),
    ]:
        summary_sheet.merge_cells(f"A{row}:D{row}")
        summary_sheet.merge_cells(f"E{row}:H{row}")
        summary_sheet[f"A{row}"] = status
        summary_sheet[f"E{row}"] = f"{days}일"
        style_range(summary_sheet, f"A{row}:H{row}", font=body_font, alignment=center)
        row += 1

    row += 1
    row = put_section_title(summary_sheet, row, "3. AWD 수행 분석 및 탄소감축 산정 기반")
    row = put_kv(summary_sheet, row, "AWD 수행 기준", "DRY 상태 이후 DRYING, FLOODED 또는 OVERFLOODED 상태로 전환되는 경우를 1회로 정의", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "AWD 수행 횟수", f"{report.total_awd_cycles}회", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "탄소감축량 산정", "향후 적용 필요", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "탄소감축량", "미산정", "A{row}:B{row}", "C{row}:H{row}")
    row += 1
    row = put_section_title(summary_sheet, row, "4. 검증 결과")
    row = put_kv(summary_sheet, row, "검증 방법", validation["validation_method"], "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "샘플 수", f"{validation['validation_sample_count']}건", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "일치 / 불일치", f"{validation['validation_match_count']}건 / {validation['validation_mismatch_count']}건", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "센서-관찰 검증 정확도", f"{validation['validation_accuracy']}%", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "AI-센서 일치", f"{validation['ai_sensor_match_count']}건", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "AI-센서 불일치", f"{validation['ai_sensor_mismatch_count']}건", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "AI-센서 판별 불가", f"{validation['ai_sensor_unknown_count']}건", "A{row}:B{row}", "C{row}:H{row}")
    row = put_kv(summary_sheet, row, "AI-센서 일치율", f"{validation['ai_sensor_accuracy']}%", "A{row}:B{row}", "C{row}:H{row}")

    row += 1
    row = put_section_title(summary_sheet, row, "5. 결론 및 향후 계획")
    conclusion = (
        f"분석 기간 동안 주요 상태는 {dominant_status}로 확인되었으며, "
        f"AWD 수행 횟수는 {report.total_awd_cycles}회로 집계되었다. "
        "공식 탄소감축량 산정은 향후 과제이며, "
        "본 보고서는 MRV 기반 데이터 관리 결과를 제시한다."
    )
    if report.total_awd_cycles == 0:
        plan = "향후 DRY 상태 이후 적절한 시점의 계획적 재관개와 현장 검증 데이터 확보가 필요하다."
    else:
        plan = (
            "향후 장기 실증 데이터 확보, 탄소감축량 산정 모델 적용, "
            "탄소배출권 제도 연계를 통해 MRV 활용 범위를 확장할 필요가 있다."
        )
    summary_sheet.merge_cells(f"A{row}:H{row + 2}")
    summary_sheet[f"A{row}"] = f"{conclusion}\n{plan}"
    style_range(summary_sheet, f"A{row}:H{row + 2}", fill=note_fill, font=body_font, alignment=left)
    summary_sheet.row_dimensions[row].height = 55

    daily_grouped = defaultdict(list)
    for s in summaries:
        daily_grouped[s.record_date].append(s)

    representative_by_date = {s.record_date: s for s in daily_summaries}
    node_ids = [node.id for node in nodes]

    flow_sheet = workbook.create_sheet(title="날짜별 흐름 데이터")
    flow_sheet.sheet_view.showGridLines = False

    flow_headers = ["날짜", "대표 평균 수위(cm)", "대표 상태"]
    for node_id in node_ids:
        flow_headers.extend([f"노드 {node_id} 수위(cm)", f"노드 {node_id} 상태"])

    flow_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(flow_headers))
    flow_sheet["A1"] = "날짜별 흐름 데이터"
    flow_sheet["A1"].font = title_font
    flow_sheet["A1"].alignment = center
    flow_sheet.row_dimensions[1].height = 30
    flow_sheet.append([])
    flow_sheet.append(flow_headers)

    for col_idx in range(1, len(flow_headers) + 1):
        cell = flow_sheet.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    flow_sheet.column_dimensions["A"].width = 16
    flow_sheet.column_dimensions["B"].width = 20
    flow_sheet.column_dimensions["C"].width = 18
    for col_idx in range(4, len(flow_headers) + 1):
        flow_sheet.column_dimensions[get_column_letter(col_idx)].width = 18

    current_row = 4
    for record_date in sorted(daily_grouped.keys()):
        node_map = {item.node_id: item for item in daily_grouped[record_date]}
        representative = representative_by_date.get(record_date)
        row_values = [
            str(record_date),
            round(float(representative.avg_inner_level), 2) if representative and representative.avg_inner_level is not None else None,
            representative.daily_status if representative else None,
        ]
        for node_id in node_ids:
            item = node_map.get(node_id)
            row_values.extend([
                round(float(item.avg_inner_level), 2) if item and item.avg_inner_level is not None else None,
                item.daily_status if item else None,
            ])

        for col_idx, value in enumerate(row_values, start=1):
            cell = flow_sheet.cell(row=current_row, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = center
            cell.border = border
        current_row += 1

    detail_sheet = workbook.create_sheet(title="노드별 상세 데이터")
    detail_sheet.sheet_view.showGridLines = False
    set_widths(detail_sheet, {"A": 16, "B": 12, "C": 18, "D": 18, "E": 45})
    detail_sheet.merge_cells("A1:E1")
    detail_sheet["A1"] = "노드별 상세 데이터"
    detail_sheet["A1"].font = title_font
    detail_sheet["A1"].alignment = center
    detail_sheet.row_dimensions[1].height = 30
    detail_sheet.append([])
    detail_sheet.append(["날짜", "노드 ID", "평균 내부 수위(cm)", "일일 상태", "검증 이미지 URL"])
    style_range(detail_sheet, "A3:E3", fill=header_fill, font=header_font, alignment=center)

    thick_side = Side(style="medium", color="404040")

    current_row = 4
    for record_date in sorted(daily_grouped.keys()):
        items = sorted(daily_grouped[record_date], key=lambda x: x.node_id)
        group_start = current_row

        for idx, item in enumerate(items):
            detail_sheet.cell(row=current_row, column=1, value=str(record_date) if idx == 0 else None)
            detail_sheet.cell(row=current_row, column=2, value=item.node_id)
            detail_sheet.cell(
                row=current_row,
                column=3,
                value=round(float(item.avg_inner_level), 2) if item.avg_inner_level is not None else None,
            )
            detail_sheet.cell(row=current_row, column=4, value=item.daily_status)
            detail_sheet.cell(row=current_row, column=5, value=item.verification_image_url)
            current_row += 1

        group_end = current_row - 1

        if group_end > group_start:
            detail_sheet.merge_cells(start_row=group_start, start_column=1, end_row=group_end, end_column=1)

        for row_idx in range(group_start, group_end + 1):
            for col_idx in range(1, 6):
                cell = detail_sheet.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.alignment = left if col_idx == 5 else center
                cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thick_side if row_idx == group_start else thin_side,
                    bottom=thick_side if row_idx == group_end else thin_side,
                )

        detail_sheet.row_dimensions[group_end].height = 22

    validation_sheet = workbook.create_sheet(title="검증 상세")
    validation_sheet.sheet_view.showGridLines = False
    set_widths(validation_sheet, {
        "A": 14, "B": 12, "C": 22, "D": 18, "E": 16,
        "F": 16, "G": 16, "H": 18, "I": 45, "J": 30
    })
    validation_sheet.merge_cells("A1:J1")
    validation_sheet["A1"] = "검증 상세"
    validation_sheet["A1"].font = title_font
    validation_sheet["A1"].alignment = center
    validation_sheet.row_dimensions[1].height = 30
    validation_sheet.append([])
    validation_sheet.append([
        "날짜",
        "노드 ID",
        "센서 기반 표면 판정",
        "사람 관찰값",
        "센서-관찰 검증",
        "센서 수위 구간",
        "AI 예측 구간",
        "AI-센서 구간 비교",
        "이미지 URL",
        "비고"
    ])
    
    style_range(validation_sheet, "A3:J3", fill=header_fill, font=header_font, alignment=center)

    if validation_rows:
        for row_data in validation_rows:
            if row_data.is_match is True:
                match_text = "일치"
            elif row_data.is_match is False:
                match_text = "불일치"
            else:
                match_text = "판별 불가"

            if row_data.ai_sensor_match is True:
                ai_sensor_text = "일치"
            elif row_data.ai_sensor_match is False:
                ai_sensor_text = "불일치"
            else:
                ai_sensor_text = "판별 불가"

            validation_sheet.append([
                str(row_data.record_date),
                row_data.node_id,
                row_data.sensor_predicted_status,
                row_data.observed_surface_status,
                match_text,
                get_sensor_ai_status_for_validation(row_data, db),
                row_data.ai_predicted_status,
                ai_sensor_text,
                row_data.image_url,
                row_data.note,
            ])
    else:
        validation_sheet.append(["검증 데이터 없음", "", "", "", "", "", "", "", "", ""])

    for row_cells in validation_sheet.iter_rows(min_row=4, max_row=validation_sheet.max_row, min_col=1, max_col=10):
        for cell in row_cells:
            cell.font = body_font
            cell.alignment = left if cell.column in (9, 10) else center
            cell.border = border

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A4"
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                if cell.value is not None and not cell.font:
                    cell.font = body_font

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"mrv_report_{report.id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("")
def delete_mrv_report(field_id: int, report_month: str, db: Session = Depends(get_db)):
    report = db.query(MrvReport).filter(
        MrvReport.field_id == field_id,
        MrvReport.report_month == report_month,
    ).first()

    if not report:
        raise HTTPException(status_code=404, detail="보고서 없음")

    db.delete(report)
    db.commit()

    return success_response(None, "MRV 보고서 삭제 성공")
